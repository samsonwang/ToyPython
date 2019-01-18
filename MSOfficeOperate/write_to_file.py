
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def write_data(fname, data):
    print('write data ...')
    wb = Workbook()
    ws = wb.active
    ws.title = fname
    r = 1 # row
    r_f = 0
    r_t = 0
    r_c = 0
    for f in data.keys():
        feature = data[f]
        print('processing ...', f)
        # merge last cell
        if r_f > 0:
            ws.merge_cells(start_row=r_f, start_column=2, end_row=r-1, end_column=2)
        r_f = r
        # feature point column
        ws.cell(column=2, row=r).value = f
        ws.cell(column=2, row=r).alignment = Alignment(wrap_text=True, vertical='center')
        for t in feature.keys():
            if r_t > 0:
                ws.merge_cells(start_row=r_t, start_column=3, end_row=r-1, end_column=3)
                ws.merge_cells(start_row=r_t, start_column=5, end_row=r-1, end_column=5)
            r_t = r
            ws.cell(column=3, row=r).value = t
            ws.cell(column=3, row=r).alignment = Alignment(wrap_text=True, vertical='center')
            test = feature[t]
            ws.cell(column=5, row=r).value=len(test)
            ws.cell(column=5, row=r).alignment = Alignment(wrap_text=True, vertical='center')
            case_count = 1
            for c in test:
                ws.cell(column=4, row=r).value = '%d、' % case_count + c
                # assign
                ws.cell(column=4, row=r).alignment = Alignment(wrap_text=True, vertical='center')
                case_count += 1
                r += 1

    ws.merge_cells(start_row=r_f, start_column=2, end_row=r-1, end_column=2)
    ws.merge_cells(start_row=r_t, start_column=3, end_row=r-1, end_column=3)
    ws.merge_cells(start_row=r_t, start_column=5, end_row=r-1, end_column=5)

    ws.column_dimensions['D'].width = 100
    wb.save(fname+'.xlsx')
