#!/usr/bin/python3

import os
import sys

from collections import defaultdict

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from write_to_file import *

#dates_dict = defaultdict(list)
g_data = dict()

g_major = 'ISCS'

def generate_case(oper, result):
    if oper.endswith('。'):
        oper = oper[0:-1]
    if oper[1] == '、':
        oper = oper[2:]
    if result and not result.endswith('。'):
        result += '。'

    case = []
    if result:
        case.append('打开中心终端；' + oper + '；' + result)
        case.append('打开车站终端；' + oper + '；' + result)
    else:
        case.append('打开中心终端；' + oper + '。')
        case.append('打开车站终端；' + oper + '。')
    return case

def load_data():
    print("load data ...")
    wb = load_workbook(filename='iscs_testcase.xlsx', read_only=True)
    ws = wb['Sheet1']

    for row in ws.rows:
        # FAS
        if row[3].value == g_major:
            # feature point
            if row[4].value not in g_data:
                g_data[row[4].value] = defaultdict(list)
            # test point
#            if row[5].value not in row[4].value:
#                g_data[row[4].value][row[5].value] = list()
            # test case
            test_case = generate_case(row[6].value, row[7].value)
            for case in test_case:
                g_data[row[4].value][row[5].value].append(case)

def main():
    g_major = 'ISCS'
    load_data()
    print(g_data)
    write_data(g_major, g_data)


if __name__ == "__main__":
    main()
