#!/usr/bin/python3

import os
import sys
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from collections import defaultdict

#dates_dict = defaultdict(list)
g_data = dict()

g_major = 'FAS'

def generate_case(oper, result):
    if oper.endswith('。'):
        oper = oper[0:-1]
    if oper[1] == '、':
        oper = oper[2:]
    if not result.endswith('。'):
        result += '。'

    case = []
    case.append('打开中心终端；' + oper + '；' + result)
    case.append('打开车站终端；' + oper + '；' + result)
    return case

def load_data():
    print("load data ...")
    wb = load_workbook(filename='iscs_testcase.xlsx', read_only=True)
    ws = wb['Sheet1']

    for row in ws.rows:
        # FAS
        if row[3].value == g_major:
            # feature point
            if row[4].value not in g_data:
                g_data[row[4].value] = defaultdict(list)
            # test point
#            if row[5].value not in row[4].value:
#                g_data[row[4].value][row[5].value] = list()
            # test case
            test_case = generate_case(row[6].value, row[7].value)
            for case in test_case:
                g_data[row[4].value][row[5].value].append(case)


def write_data():
    print('write data ...')
    wb = Workbook()
    ws = wb.active
    ws.title = g_major
    r = 1 # row
    r_f = 0
    r_t = 0
    r_c = 0
    for f in g_data.keys():
        feature = g_data[f]
        print('processing ...', f)
        # merge last cell
        if r_f > 0:
            ws.merge_cells(start_row=r_f, start_column=2, end_row=r-1, end_column=2)
        r_f = r
        # feature point column
        ws.cell(column=2, row=r).value = f
        ws.cell(column=2, row=r).alignment = Alignment(wrap_text=True, vertical='center')
        for t in feature.keys():
            if r_t > 0:
                ws.merge_cells(start_row=r_t, start_column=3, end_row=r-1, end_column=3)
                ws.merge_cells(start_row=r_t, start_column=5, end_row=r-1, end_column=5)
            r_t = r
            ws.cell(column=3, row=r).value = t
            ws.cell(column=3, row=r).alignment = Alignment(wrap_text=True, vertical='center')
            test = feature[t]
            ws.cell(column=5, row=r).value=len(test)
            ws.cell(column=5, row=r).alignment = Alignment(wrap_text=True, vertical='center')
            case_count = 1
            for c in test:
                ws.cell(column=4, row=r).value = '%d、' % case_count + c
                # assign
                ws.cell(column=4, row=r).alignment = Alignment(wrap_text=True, vertical='center')
                case_count += 1
                r += 1

    ws.column_dimensions['D'].width = 100

#    cell.alignment = wrap_alignment

    wb.save(g_major+'.xlsx')

def main():
    g_major = 'FAS'
    load_data()
    print(g_data)
    write_data()


if __name__ == "__main__":
    main()
